import logging
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__package__)


def excel(data: List[Dict[str, str]]):
    logger.info('Crate xlsx file')
    wb = Workbook()
    ws = wb.active
    logger.debug('Create Title')
    header = ('#', 'Title', 'Link')
    for col, text in enumerate(header, 1):
        title = ws.cell(row=1, column=col)
        title.value = text
        title.fill = PatternFill("solid", fgColor="00CCCCFF")
        title.font = Font(bold=True)

    logger.debug('Create Body')
    for count, line in enumerate(data, 1):
        row = count + 1
        cell = ws.cell(row=row, column=1)
        cell.value = count
        cell = ws.cell(row=row, column=2)
        cell.value = line['name']
        cell = ws.cell(row=row, column=3)
        cell.value = line['id']
        cell.hyperlink = line['url']
        cell.font = Font(color="000000FF")
    ws.column_dimensions['B'].width = max(map(lambda x: len(x['name']), data))
    ws.column_dimensions['C'].width = max(map(lambda x: len(x['id']), data))

    file_name = f'{datetime.now().strftime("%Y_%m_%d_%H_%M_%S")}.xlsx'
    logger.info(f'Save file: {file_name}')
    wb.save(file_name)
